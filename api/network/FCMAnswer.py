import time

from openpyxl import Workbook

from api.network.utils.DataMapHelper import DataMapHelper
from api.network.utils.FCMComputer import FCMComputer
from api.network.IAnswer import IAnswer

import openpyxl


class FCMAnswer(IAnswer):
    fcm_answer = None

    def __init__(self):
        super().__init__()
        self.data_map_helper = DataMapHelper.getInstance()
        self.fcm_computer = FCMComputer()

    def getAllSegmentName(self):
        return self.data_map_helper.getAllSegmentName()

    def getSegmentExplain(self, name):
        return self.data_map_helper.getSegmentExplain(name)

    def getCompanyLabel(self, data_dic, company_name):
        return self.fcm_computer.compute(data_dic, company_name)

    def getSegmentBelong(self, name):
        return self.data_map_helper.getBelongIndex(name)

    def getBelongMap(self, index):
        return self.data_map_helper.getBelongFromIndex(index)

    def get_segment_input_content(self, segment_name):
        return self.data_map_helper.get_segment_input_content(segment_name)

    def solve_unaccept_value(self, segment, value):
        return self.data_map_helper.solve_unaccept_value(segment, value)

    def alreadyInDataBase(self, company_name):
        return self.fcm_computer.alreadyInDataBase(company_name)

    def getCompanyLabelFromExecel(self, execel):

        execel = openpyxl.load_workbook(execel)
        if not isinstance(execel, Workbook):
            raise ValueError("输入的参数不对")

        label_dic = {}
        for sheet in execel:
            for row, _ in enumerate(sheet.rows):
                if row == 0:
                    continue
                row += 1
                data_dic = {}
                company_name = sheet.cell(row=row, column=1).value
                for col, _ in enumerate(sheet.columns):
                    if col == 0:
                        continue
                    col += 1
                    try:
                        val = sheet.cell(row=row, column=col).value
                        data_dic[sheet.cell(row=1, column=col).value] = float(val) if val != '' and val is not None else None
                    except:
                        raise ValueError("execel文件格式错误")
                # try:
                label = self.getCompanyLabel(data_dic, company_name)
                label_dic[company_name] = label
                # except:
                #     raise ValueError("execel文件格式错误")
        return label_dic

    @staticmethod
    def getInstance():
        if FCMAnswer.fcm_answer is None:
            FCMAnswer.fcm_answer = FCMAnswer()
        return FCMAnswer.fcm_answer


if __name__ == "__main__":
    """
    测试
    """
    instance = FCMAnswer.getInstance()
    beg = time.time()
    instance.getCompanyLabelFromExecel('C:\\Users\\bullypaulo\\Desktop\\2020服务外包大赛\\批量操作测试.xlsx')
    end = time.time()
    print("耗时：" + str(end - beg))
    pass
